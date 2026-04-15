"""
==============================================================================
CLMRS HOUSEHOLD PROFILING — SCRIPT HQ APPROBATION/REJET
==============================================================================
Projet   : CIV - CLMRS Household Profiling 25-26
Rôle     : HEADQUARTERS (HQ) — Approbation/Rejet niveau quartier général


RÈGLES DE VALIDATION (exhaustives — rien d'autre n'est appliqué) :

  1. PROGRESSION / COMPLÉTUDE
       a. Rejet si ForInterviewer == 0  (enquête vide ou corrompue)
       b. Rejet si ForInterviewer < 50  (sections sautées anormalement)

  2. TEMPOREL / GPS
       a. Rejet si durée entre `date` et `end_interview` < 45 minutes
       b. Rejet si `date` ou `end_interview` manquant
       c. Rejet si coordonnées GPS (`hhgps`) absentes
       d. Rejet si précision GPS > 20 mètres

  3. CONSENTEMENT / FRAUDE
       a. Rejet si `consent` ne commence pas par "Oui" (insensible à la casse)
       b. Rejet si liste préchargée présente ET tous les enfants ont interview_child==1
          (même si de nouveaux enfants ont été déclarés)
       c. Rejet si `children`=Oui MAIS 0 dossier enfant rempli
       d. Rejet si  il y a liste d'enfants pré-rempli mais children=Non (contradiction)
       e. Rejet si child_year_birth_new absent ou invalide pour un enfant interviewé
       f. Rejet si new_child_year_birth absent ou invalide pour un nouvel enfant déclaré
       g. Rejet si new_child_list ne contient que des chiffres (sans préchargé)

  4. FORMATAGE
       Seuls les champs de saisie texte libre sont contrôlés (TEXT_VARS et
       LIST_TEXT_VARS). Les questions à propositions de réponses sont exclues.
       a. Rejet si un champ de saisie libre contient des lettres minuscules
       b. Rejet si un champ de saisie libre contient des caractères accentués

WORKFLOW REJET/APPROBATION :
  - Approbation          : /hqapprove
  - Rejet Completed      : /reject   avec params comment (visible sur tablette)
  - Rejet ApprovedBySup  : /unapprove puis /reject avec params comment
  - Rejet RejectedBySup  : /hqreject avec params comment

PROGRESSION :
  - Numérateur  : Answered   (questions répondues)
  - Dénominateur: ForInterviewer (questions activées visibles enquêteur)
  - Seuil rejet : ForInterviewer < 50

USAGE :
  CLMRS_active_script.py               → Production HQ
  CLMRS_active_script.py --test        → Lecture seule
  CLMRS_active_script.py --diagnostic  → Inspecte 5 enquêtes
  CLMRS_active_script.py --limit=20    → Limite à 20 enquêtes
  CLMRS_active_script.py --reset       → Efface le checkpoint

INSTALLATION :
  pip install requests tenacity openpyxl
==============================================================================
"""

import json
import logging
import os
import re
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import requests
from tenacity import (
    RetryError,
    before_sleep_log,
    retry,
    retry_if_exception_type,
    stop_after_attempt,
    wait_exponential,
)

# ==============================================================================
# 1. CONFIGURATION
# ==============================================================================

CONFIG = {
    # ── Connexion ──────────────────────────────────────────────────────────────
    "hq_url":                  "https://mysurvey.touton.com",
    "api_user":                "",
    "api_password":            "",
    "workspace":               "primary",
    "questionnaire_id":        "",
    "questionnaire_version":   ,

    # ── Statuts récupérés ─────────────────────────────────────────────────────
    "status_to_fetch": [
        "Completed",
        "ApprovedBySupervisor",
        "RejectedBySupervisor",
    ],

    # ── Pagination / délais ───────────────────────────────────────────────────
    "page_size":             40,
    "sleep_between_calls":   0.3,

    # ── Fichiers de sortie ────────────────────────────────────────────────────
    "log_file":              "validation_hq.log",
    "export_report":         True,
    "report_file":           "rapport_validation_hq.xlsx",
    "checkpoint_file":       "checkpoint_hq.json",
    "diagnostic_file":       "diagnostic_api.log",

    # ── Règle 1 : Progression ─────────────────────────────────────────────────
    # Rejet si ForInterviewer (questions activées) < 50
    "min_total_questions":   50,

    # ── Règle 2 : Temporel / GPS ──────────────────────────────────────────────
    "min_duration_minutes":  45,
    "gps_accuracy_max_m":    20.0,

    # ── Réseau — tenacity ─────────────────────────────────────────────────────
    "retry_attempts":        4,
    "retry_wait_min":        2,
    "retry_wait_max":        15,
}

# ==============================================================================
# 2. LOGGING
# ==============================================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(CONFIG["log_file"], encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger(__name__)

diag_logger = logging.getLogger("diagnostic")
_dh = logging.FileHandler(CONFIG["diagnostic_file"], encoding="utf-8")
_dh.setFormatter(logging.Formatter("%(asctime)s - %(message)s"))
diag_logger.addHandler(_dh)
diag_logger.setLevel(logging.DEBUG)
diag_logger.propagate = False

# ==============================================================================
# 3. UTILITAIRE — safe_get
# ==============================================================================

def safe_get(data: Any, *keys, default: Any = None) -> Any:
    """Navigation défensive dans les dicts/listes imbriqués."""
    try:
        for key in keys:
            if data is None:
                return default
            data = data[key]
        return data if data is not None else default
    except (KeyError, TypeError, IndexError, AttributeError):
        return default

# ==============================================================================
# 4. CHECKPOINT
# ==============================================================================

def load_checkpoint() -> set:
    path = CONFIG["checkpoint_file"]
    if os.path.exists(path):
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
            ids = set(data) if isinstance(data, list) else set()
            logger.info(f"Checkpoint chargé : {len(ids)} enquête(s) déjà traitée(s).")
            return ids
        except Exception as e:
            logger.warning(f"Checkpoint illisible ({e}) — démarrage à zéro.")
    return set()


def save_checkpoint(done_ids: set) -> None:
    try:
        with open(CONFIG["checkpoint_file"], "w", encoding="utf-8") as f:
            json.dump(list(done_ids), f)
    except Exception as e:
        logger.warning(f"Impossible de sauvegarder le checkpoint : {e}")


def clear_checkpoint() -> None:
    path = CONFIG["checkpoint_file"]
    if os.path.exists(path):
        try:
            os.remove(path)
            logger.info("Checkpoint effacé.")
        except Exception as e:
            logger.warning(f"Impossible de supprimer le checkpoint : {e}")

# ==============================================================================
# 5. EXPORT EXCEL
# ==============================================================================

def export_excel(results: list, filepath: str) -> None:
    rows = [
        {
            "ID":             r.interview_id,
            "Clé":            r.interview_key,
            "Enquêteur":      r.enumerator,
            "Statut initial": r.statut_initial,
            "Décision HQ":    r.decision,
            "Progression":    r.progress_str,
            "Réponses":       r.answered_count,
            "Total activées": r.total_questions,
            "Enfants":        r.children_count,
            "Nb anomalies":   len(r.errors),
            "Commentaire":    r.comment.replace("\n", " | "),
        }
        for r in results
    ]

    try:
        import pandas as pd
        pd.DataFrame(rows).to_excel(filepath, index=False)
        logger.info(f"Rapport Excel exporté via pandas : {filepath}")
        return
    except ImportError:
        logger.info("pandas absent — bascule sur openpyxl.")
    except Exception as e:
        logger.warning(f"pandas erreur ({e}) — bascule sur openpyxl.")

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Validation HQ"
        headers = list(rows[0].keys()) if rows else []

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="1A3A5C")
            cell.alignment = Alignment(horizontal="center")

        fill_ok = PatternFill("solid", fgColor="C6EFCE")
        fill_ko = PatternFill("solid", fgColor="FFC7CE")

        for ri, row in enumerate(rows, 2):
            fill = fill_ok if row.get("Décision HQ") == "HQ_APPROVE" else fill_ko
            for ci, h in enumerate(headers, 1):
                cell      = ws.cell(row=ri, column=ci, value=row[h])
                cell.fill = fill

        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        wb.save(filepath)
        logger.info(f"Rapport Excel exporté via openpyxl : {filepath}")
    except ImportError:
        logger.warning("openpyxl absent. Installez : pip install openpyxl")
    except Exception as e:
        logger.error(f"Export Excel impossible : {e}")

# ==============================================================================
# 6. MODÈLE
# ==============================================================================

@dataclass
class ValidationResult:
    interview_id:    str
    enumerator:      str
    interview_key:   str  = ""
    statut_initial:  str  = ""
    decision:        str  = "HQ_APPROVE"
    errors:          list = field(default_factory=list)
    comment:         str  = ""
    progress_str:    str  = "0/0"
    answered_count:  int  = 0
    total_questions: int  = 0
    children_count:  int  = 0

# ==============================================================================
# 7. CLIENT API HQ
# ==============================================================================

_RETRY_EXC = (
    requests.exceptions.Timeout,
    requests.exceptions.ConnectionError,
    requests.exceptions.ChunkedEncodingError,
)


class SurveySolutionsHQClient:
    """Client API Survey Solutions avec droits Headquarters."""

    def __init__(self):
        self.base = CONFIG["hq_url"].rstrip("/")
        self.ws   = CONFIG["workspace"]
        self.auth = (CONFIG["api_user"], CONFIG["api_password"])

    def _url(self, endpoint: str) -> str:
        return f"{self.base}/{self.ws}/api/v1/{endpoint}"

    @retry(
        stop=stop_after_attempt(CONFIG["retry_attempts"]),
        wait=wait_exponential(multiplier=1,
                              min=CONFIG["retry_wait_min"],
                              max=CONFIG["retry_wait_max"]),
        retry=retry_if_exception_type(_RETRY_EXC),
        before_sleep=before_sleep_log(logger, logging.WARNING),
        reraise=False,
    )
    def _raw_get(self, url: str, params=None) -> Optional[requests.Response]:
        return requests.get(url, auth=self.auth, params=params, timeout=30)

    @retry(
        stop=stop_after_attempt(CONFIG["retry_attempts"]),
        wait=wait_exponential(multiplier=1,
                              min=CONFIG["retry_wait_min"],
                              max=CONFIG["retry_wait_max"]),
        retry=retry_if_exception_type(_RETRY_EXC),
        before_sleep=before_sleep_log(logger, logging.WARNING),
        reraise=False,
    )
    def _raw_patch(self, url: str, params=None) -> Optional[requests.Response]:
        return requests.patch(url, auth=self.auth, params=params, timeout=30)

    def _get(self, endpoint: str, params=None) -> Optional[Dict]:
        url = self._url(endpoint)
        try:
            r = self._raw_get(url, params)
            if r is None:
                logger.warning(f"GET {endpoint} : toutes les tentatives échouées.")
                return None
            if r.status_code == 200:
                return r.json()
            logger.warning(f"GET {endpoint} → HTTP {r.status_code} | {r.text[:300]}")
        except RetryError:
            logger.warning(f"GET {endpoint} : RetryError.")
        except Exception as e:
            logger.warning(f"GET {endpoint} : {type(e).__name__} — {e}")
        return None

    def _patch(self, endpoint: str, params=None) -> Tuple[bool, str]:
        """PATCH avec paramètres en query string (comment visible sur tablette)."""
        url = self._url(endpoint)
        try:
            r = self._raw_patch(url, params)
            if r is None:
                return False, "Toutes les tentatives réseau ont échoué."
            if r.status_code == 200:
                return True, ""
            msg = f"HTTP {r.status_code} — {r.text[:300]}"
            logger.warning(f"PATCH {endpoint} → {msg}")
            return False, msg
        except RetryError:
            return False, "RetryError."
        except Exception as e:
            return False, f"{type(e).__name__}: {e}"

    def test_connection(self) -> bool:
        logger.info(f"Test connexion HQ → {self.base} ...")
        if self._get("questionnaires") is not None:
            logger.info("Connexion HQ réussie.")
            return True
        logger.error(
            "Connexion HQ échouée. Vérifiez : URL, identifiants, workspace, "
            "et que le compte a bien le rôle Headquarters."
        )
        return False

    def get_interviews(self, limit: Optional[int] = None) -> List[Dict]:
        toutes = []
        for status in CONFIG["status_to_fetch"]:
            page = 1
            while True:
                params = {
                    "status":               status,
                    "questionnaireId":      CONFIG["questionnaire_id"],
                    "questionnaireVersion": CONFIG["questionnaire_version"],
                    "page":                 page,
                    "pageSize":             CONFIG["page_size"],
                }
                data       = self._get("interviews", params=params)
                interviews = safe_get(data, "Interviews", default=[])
                if not interviews:
                    break
                toutes.extend(interviews)
                logger.info(f"  [{status}] Page {page} : {len(interviews)} enquête(s)")
                if len(interviews) < CONFIG["page_size"]:
                    break
                page += 1
                time.sleep(CONFIG["sleep_between_calls"])
        return toutes[:limit] if limit else toutes

    def get_interview_details(self, iid: str) -> Optional[Dict]:
        return self._get(f"interviews/{iid}")

    def get_interview_stats(self, iid: str) -> Optional[Dict]:
        for ep in [f"interviews/{iid}/stats", f"interviews/{iid}/statistics"]:
            data = self._get(ep)
            if data and isinstance(data, dict):
                return data
        return None

    def hq_approve(self, iid: str, comment: str) -> Tuple[bool, str]:
        """Approbation HQ directe."""
        return self._patch(
            f"interviews/{iid}/hqapprove",
            params={"comment": comment[:1500]},
        )

    def supervisor_unapprove(self, iid: str) -> Tuple[bool, str]:
        """Repasse ApprovedBySupervisor → Completed pour pouvoir faire /reject."""
        return self._patch(f"interviews/{iid}/unapprove")

    def supervisor_reject(self, iid: str, comment: str) -> Tuple[bool, str]:
        """
        Rejet pour Completed et ApprovedBySupervisor.
        Le commentaire en query param est visible sur la tablette.
        """
        return self._patch(
            f"interviews/{iid}/reject",
            params={"comment": comment[:1500]},
        )

    def hq_reject(self, iid: str, comment: str) -> Tuple[bool, str]:
        """
        Rejet pour RejectedBySupervisor (seul statut accepté par /hqreject).
        Le commentaire en query param est visible sur la tablette.
        """
        return self._patch(
            f"interviews/{iid}/hqreject",
            params={"comment": comment[:1500]},
        )

    def run_diagnostic(self, entretiens: List[Dict], nb: int = 5) -> None:
        logger.info("=" * 60)
        logger.info(f"MODE DIAGNOSTIC — {min(nb, len(entretiens))} enquête(s)")
        logger.info("=" * 60)
        for inter in entretiens[:nb]:
            iid    = safe_get(inter, "InterviewId", default="?")
            status = safe_get(inter, "Status",      default="?")
            logger.info(f"\n── Enquête : {iid}  (statut={status})")
            for suffix in ["", "/stats"]:
                url = self._url(f"interviews/{iid}{suffix}")
                try:
                    r   = requests.get(url, auth=self.auth, timeout=30)
                    msg = f"  [GET {suffix or '/details'}]  HTTP {r.status_code}  ({len(r.content)} octets)"
                    if r.status_code == 200 and suffix == "/stats":
                        msg += f"\n    Contenu stats : {json.dumps(r.json(), ensure_ascii=False)}"
                    elif r.status_code != 200:
                        msg += f"\n    Corps : {r.text[:400]}"
                    logger.info(msg)
                    diag_logger.info(msg)
                except Exception as e:
                    logger.warning(f"  [GET {suffix}] EXCEPTION : {e}")
        logger.info("\nCodes HTTP :")
        logger.info("  200 → OK | 401 → Identifiants | 403 → Droits insuffisants")
        logger.info("  404 → Endpoint introuvable | 406 → Statut incompatible | 500 → Erreur serveur")
        logger.info(f"\nRapport complet → {CONFIG['diagnostic_file']}")
        logger.info("=" * 60)

# ==============================================================================
# 8. EXTRACTEUR DE VARIABLES
# ==============================================================================

class HouseholdVariableExtractor:
    """Parse les réponses JSON de l'API Survey Solutions."""

    def __init__(self, details: dict):
        self._vars:    Dict[str, Any]  = {}
        self._rosters: Dict[int, Dict] = {}

        for ans in safe_get(details, "Answers", default=[]):
            try:
                var_name = safe_get(ans, "VariableName")
                val      = safe_get(ans, "Answer")
                if not var_name:
                    continue
                rv = safe_get(ans, "QuestionId", "RosterVector", default=[])
                if rv:
                    row_index = rv[0]
                    if row_index not in self._rosters:
                        self._rosters[row_index] = {}
                    self._rosters[row_index][var_name] = val
                else:
                    self._vars[var_name] = val
            except Exception:
                continue

    def get(self, name: str, default: Any = None) -> Any:
        return self._vars.get(name, default)

    def get_int(self, name: str, default: int = -1) -> int:
        try:
            v = self._vars.get(name)
            return int(v) if v is not None else default
        except Exception:
            return default

    def get_bool_from_str(
        self, name: str,
        true_values: tuple = ("oui", "yes", "1", "true"),
    ) -> Optional[bool]:
        v = self._vars.get(name)
        if v is None:
            return None
        return str(v).strip().lower() in true_values

    def get_roster_rows(self) -> List[Dict]:
        return list(self._rosters.values())

# ==============================================================================
# 9. VARIABLES DE SAISIE LIBRE
# ==============================================================================

TEXT_VARS = {
    "consent_why_no", "exact_name_entered", "ethnicity", "country_origin_other",
    "farmer_status_other", "name_owner", "first_name_owner", "ethnicity_owner",
    "country_origin_owner_other", "worker_agreement_other", "refusal_worker_other",
    "adult_family_rlshp_other", "adult_country_origin_other", "adult_main_work_other",
    "adult_answer_other", "duplicates_child_name", "child_surname",
    "children_leaving_hh_other", "child_not_avail_other", "hh_headAutre",
    "child_parentOther", "child_family_decidedOther", "child_family_withother",
    "child_father_countryOther", "child_mother_countryOther", "child_school_name",
    "child_schl_left_whyOther", "child_why_no_schoolother", "other_reason_miss",
    "child_work_whoother", "child_work_whyother", "other_help_child",
    "immediate_danger_why", "child_remediation120ther", "child_remediation14a",
    "child_remediation16a", "new_who_ans_other", "new_child_not_availOther",
    "new_hh_headAutre", "new_child_parentother", "new_child_family_decidedother",
    "new_child_family_withother", "new_child_father_countryOther",
    "new_child_mother_countryOther", "new_child_school_name",
    "new_child_sch1_left_whyOther", "new_child_why_no_schoolOther",
    "new_other_reason_miss", "new_child_work_whoother", "new_child_work_whyOther",
    "new_other_help_child", "new_immediate_danger_why", "new_child_remediation120ther",
    "new_child_remediation14", "new_child_remediation16", "child_remediationll_other",
    "child_remediation12", "child_remediation14_other", "child_remediation15",
    "child_remediation16_other", "remediation17_other", "tools_other", "feedback_enum",
}

LIST_TEXT_VARS = {"new_child_list", "list_adults"}

_RE_LOWERCASE = re.compile(r"[a-z]")
_RE_ACCENT    = re.compile(r"[éèêëàâäîïôöûüùçÉÈÊËÀÂÄÎÏÔÖÛÜÙÇ]")

# ==============================================================================
# 10. HELPERS interview_child / duplicates / date de naissance
# ==============================================================================

def _is_interview_child_esquive(val) -> bool:
    """
    interview_child == 1 / "Oui" → enfant ESQUIVÉ.
    L'enquêteur prétend avoir déjà interviewé cet enfant → roster non ouvert.
    """
    if val is None:
        return False
    s = str(val).strip().lower()
    if s in ("oui", "yes", "1", "true"):
        return True
    try:
        return int(val) == 1
    except (TypeError, ValueError):
        return False


def _is_interview_child_interviewe(val) -> bool:
    """
    interview_child == 2 / "Non" → enfant INTERVIEWÉ.
    L'enquêteur dit que l'enfant n'a pas encore été interviewé → roster ouvert.
    """
    if val is None:
        return False
    s = str(val).strip().lower()
    if s in ("non", "no", "2", "false"):
        return True
    try:
        return int(val) == 2
    except (TypeError, ValueError):
        return False


def _is_duplicate(val) -> bool:
    """duplicates == 1 / "Oui" → enfant doublon."""
    if val is None:
        return False
    s = str(val).strip().lower()
    if s in ("oui", "yes", "1", "true"):
        return True
    try:
        return int(val) == 1
    except (TypeError, ValueError):
        return False


def _parse_birth_year(val) -> tuple:
    """
    Parse une date de naissance.
    Retourne (année:int, None) si valide, (None, raison:str) si invalide/absent.
    """
    if val is None:
        return None, "absent"
    try:
        year = datetime.fromisoformat(str(val).replace("Z", "")).year
        return year, None
    except Exception:
        return None, f"illisible ('{val}')"

# ==============================================================================
# 11. MOTEUR DE VALIDATION HQ
# ==============================================================================

class HQInterviewValidator:

    def __init__(
        self,
        interview_meta: dict,
        details: Optional[dict],
        stats: Optional[dict] = None,
    ):
        self.ev     = HouseholdVariableExtractor(details) if details else None
        self.stats  = stats
        self.result = ValidationResult(
            interview_id   = safe_get(interview_meta, "InterviewId",     default=""),
            interview_key  = safe_get(interview_meta, "Key",             default=""),
            enumerator     = safe_get(interview_meta, "ResponsibleName", default="Inconnu"),
            statut_initial = safe_get(interview_meta, "Status",          default=""),
        )

    def validate(self) -> ValidationResult:
        self._check_progression()

        if self.ev is None:
            self.result.errors.append(
                "Impossible d'extraire les données JSON de l'enquête "
                "(détails inaccessibles)."
            )
            self._build_comment()
            return self.result

        self._check_temporal_gps()
        self._check_consent_fraud()
        self._check_formatting()
        self._build_comment()
        return self.result

    # ── BLOC 1 — Progression ─────────────────────────────────────────────────

    def _check_progression(self) -> None:
        if not self.stats:
            self.result.errors.append(
                "Statistiques de progression inaccessibles (endpoint /stats indisponible)."
            )
            return

        # Answered  = questions effectivement répondues
        # ForInterviewer = questions activées visibles par l'enquêteur (= dénominateur)
        answered       = self.stats.get("Answered", 0)
        for_interviewer = self.stats.get("ForInterviewer", 0)

        try:
            answered        = int(answered)
            for_interviewer = int(for_interviewer)
        except (TypeError, ValueError):
            answered        = 0
            for_interviewer = 0

        self.result.answered_count  = answered
        self.result.total_questions = for_interviewer
        self.result.progress_str    = f"{answered}/{for_interviewer}"

        if for_interviewer == 0:
            self.result.errors.append(
                "REJET — Progression 0/0 : enquête vide ou corrompue."
            )
            return

        if for_interviewer < CONFIG["min_total_questions"]:
            self.result.errors.append(
                f"REJET — Progression trop faible : {for_interviewer} question(s) "
                f"activées (minimum requis : {CONFIG['min_total_questions']}). "
                "Des sections entières ont été sautées de façon anormale."
            )

    # ── BLOC 2 — Temporel / GPS ───────────────────────────────────────────────

    def _check_temporal_gps(self) -> None:
        d_start = self.ev.get("date")
        d_end   = self.ev.get("end_interview")

        if not d_start or not d_end:
            self.result.errors.append(
                "REJET — Horodatage manquant : "
                f"date={'présente' if d_start else 'ABSENTE'}, "
                f"end_interview={'présente' if d_end else 'ABSENTE'}."
            )
        else:
            try:
                start = datetime.fromisoformat(str(d_start).replace("Z", ""))
                end   = datetime.fromisoformat(str(d_end).replace("Z", ""))
                dur   = (end - start).total_seconds() / 60
                if dur < CONFIG["min_duration_minutes"]:
                    self.result.errors.append(
                        f"REJET — Durée insuffisante : {dur:.1f} min "
                        f"(minimum : {CONFIG['min_duration_minutes']} min)."
                    )
            except Exception as e:
                logger.debug(f"Durée non calculable pour {self.result.interview_id} : {e}")

        gps = self.ev.get("hhgps")
        if not gps:
            self.result.errors.append(
                "REJET — Coordonnées GPS du ménage (hhgps) absentes."
            )
        elif isinstance(gps, list) and len(gps) >= 3:
            accuracy = gps[2]
            if accuracy is None:
                self.result.errors.append(
                    "REJET — Précision GPS non renseignée (valeur nulle)."
                )
            else:
                try:
                    acc_float = float(accuracy)
                    if acc_float > CONFIG["gps_accuracy_max_m"]:
                        self.result.errors.append(
                            f"REJET — Précision GPS insuffisante : {acc_float} m "
                            f"(maximum autorisé : {CONFIG['gps_accuracy_max_m']} m)."
                        )
                except (TypeError, ValueError):
                    self.result.errors.append(
                        f"REJET — Précision GPS illisible : '{accuracy}'."
                    )

    # ── BLOC 3 — Consentement / Fraude ───────────────────────────────────────

    def _check_consent_fraud(self) -> None:
        # Règle 3a — consentement
        consent_val = self.ev.get("consent")
        if consent_val is not None:
            if not str(consent_val).strip().lower().startswith("oui"):
                self.result.errors.append(
                    f"REJET — Consentement refusé ou invalide "
                    f"(consent = '{consent_val}'). "
                    "La valeur doit commencer par 'Oui'."
                )

        # Nouvelle règle — enfants préchargés mais children=Non
        old_child_list_val = self.ev.get("old_child_list")
        has_preloaded = (
            isinstance(old_child_list_val, str)
            and old_child_list_val.strip() != ""
        )
        children_declared = self.ev.get_bool_from_str("children")
        if has_preloaded and children_declared is False:
            self.result.errors.append(
                "REJET — Contradiction : une liste d'enfants préchargée est présente "
                f"('{old_child_list_val}') mais le producteur déclare AUCUN enfant "
                "(children = Non/2). Rectifiez la réponse à la question children."
            )

        children_list, fraud_errors = self._extract_and_count_children()
        self.result.children_count = len(children_list)

        for err in fraud_errors:
            self.result.errors.append(err)

        # Règle 3c — contradiction children=Oui / 0 dossier
        if children_declared is True and self.result.children_count == 0 and not fraud_errors:
            self.result.errors.append(
                "REJET — Contradiction : le producteur déclare la présence d'enfants "
                "(children = Oui) mais AUCUN dossier enfant n'a été rempli."
            )

    def _extract_and_count_children(self) -> Tuple[List[Dict], List[str]]:
        """
        Parcourt le roster et applique les règles 3b à 3f.
        Convention interview_child :
          1 / "Oui" → ESQUIVÉ  (enquêteur prétend avoir déjà interviewé)
          2 / "Non" → INTERVIEWÉ (roster ouvert)
        """
        children     = []
        fraud_errors = []

        old_child_list_val = self.ev.get("old_child_list")
        has_preloaded = (
            isinstance(old_child_list_val, str)
            and old_child_list_val.strip() != ""
        )

        def _safe_int(val, default=-1):
            try:
                return int(val) if val is not None else default
            except Exception:
                return default

        def _safe_tasks(raw):
            if not raw:
                return []
            if not isinstance(raw, list):
                raw = [raw]
            return [int(t) for t in raw if str(t).isdigit()]

        roster_rows = self.ev.get_roster_rows()

        # ── Parcours du roster ───────────────────────────────────────────────
        for row in roster_rows:
            try:
                ic_val = row.get("interview_child")

                # ── CAS 1 : enfant INTERVIEWÉ (interview_child == 2 / "Non") ──
                if _is_interview_child_interviewe(ic_val):
                    # Règle 3d — date de naissance présente et valide
                    birth_val  = row.get("child_year_birth_new")
                    birth_year, birth_err = _parse_birth_year(birth_val)
                    if birth_err:
                        fraud_errors.append(
                            f"REJET — Date de naissance manquante ou invalide "
                            f"pour un enfant interviewé "
                            f"(child_year_birth_new {birth_err})."
                        )
                    age = (datetime.now().year - birth_year if birth_year else -1)
                    tasks = _safe_tasks(row.get("task_7d")) + _safe_tasks(row.get("task_12m"))
                    children.append({"age": age, "tasks": tasks, "source": "ancien"})

                # ── CAS 2 : enfant ESQUIVÉ (interview_child == 1 / "Oui") ──
                elif _is_interview_child_esquive(ic_val):
                    # Règle 3b — vérifier duplicates (child_year_birth_new non requis ici)
                    dup_val = row.get("duplicates")
                    if not _is_duplicate(dup_val):
                        # duplicates == 2 / "Non" → pas de doublon déclaré → rejet
                        fraud_errors.append(
                            "REJET — Enfant préchargé esquivé sans justification "
                            "(duplicates = Non/2 : l'enfant n'est pas déclaré "
                            "comme doublon). Interviewez l'enfant ou déclarez le doublon."
                        )
                    else:
                        # duplicates == 1 / "Oui" → doublon déclaré
                        # → duplicates_child_name doit être rempli
                        dup_name = row.get("duplicates_child_name")
                        if not dup_name or str(dup_name).strip() == "":
                            fraud_errors.append(
                                "REJET — Enfant préchargé déclaré doublon "
                                "mais duplicates_child_name est vide. "
                                "Sélectionnez le nom de l'enfant doublon."
                            )
                    # Enfant esquivé → ne pas compter

                # ── CAS 3 : nouvel enfant (pas de interview_child ou valeur inconnue) ──
                else:
                    new_child_list_val = row.get("new_child_list")
                    new_gender         = row.get("new_child_gender")
                    new_year           = row.get("new_child_year_birth")

                    if new_gender is not None or new_year is not None:
                        tasks = (_safe_tasks(row.get("new_task_7d"))
                                 + _safe_tasks(row.get("new_task_12m")))

                        # Règle 3e — date de naissance nouvel enfant
                        birth_year, birth_err = _parse_birth_year(new_year)
                        if birth_err:
                            fraud_errors.append(
                                f"REJET — Date de naissance manquante ou invalide "
                                f"pour un nouvel enfant "
                                f"(new_child_year_birth {birth_err})."
                            )
                        age_from_birth = (datetime.now().year - birth_year
                                          if birth_year else -1)

                        children.append({
                            "age":    age_from_birth,
                            "tasks":  tasks,
                            "source": "nouveau",
                        })

                        # Règle 3f — nom invalide (chiffres uniquement), sans préchargé
                        if not has_preloaded and new_child_list_val is not None:
                            name_str = str(new_child_list_val).strip()
                            if name_str.isdigit():
                                fraud_errors.append(
                                    f"REJET — Nom d'enfant invalide : "
                                    f"new_child_list = '{name_str}' "
                                    "ne contient que des chiffres. "
                                    "Saisissez le vrai nom."
                                )

            except Exception:
                continue

        return children, fraud_errors

    # ── BLOC 4 — Formatage ────────────────────────────────────────────────────

    def _check_formatting(self) -> None:
        def _audit_text(var_name: str, val: str, context: str = "") -> None:
            if not isinstance(val, str) or not val.strip():
                return
            prefix = f"[{context}] " if context else ""
            if _RE_LOWERCASE.search(val):
                self.result.errors.append(
                    f"REJET — Formatage {prefix}({var_name}) : "
                    f"la valeur '{val}' contient des lettres minuscules. "
                    "Saisie obligatoire en MAJUSCULES."
                )
            if _RE_ACCENT.search(val):
                self.result.errors.append(
                    f"REJET — Formatage {prefix}({var_name}) : "
                    f"la valeur '{val}' contient des caractères accentués. "
                    "Aucun accent autorisé."
                )

        for var_name, value in self.ev._vars.items():
            if var_name in TEXT_VARS:
                _audit_text(var_name, str(value) if value is not None else "")
            elif var_name in LIST_TEXT_VARS:
                items = value if isinstance(value, list) else ([value] if value else [])
                for idx, item in enumerate(items, 1):
                    texte = (
                        safe_get(item, "Text", default=str(item))
                        if isinstance(item, dict) else str(item)
                    )
                    _audit_text(var_name, texte, context=f"{var_name} element {idx}")

        for i, row in enumerate(self.ev.get_roster_rows(), 1):
            for var_name, value in row.items():
                if var_name in TEXT_VARS:
                    _audit_text(
                        var_name,
                        str(value) if value is not None else "",
                        context=f"Roster ligne {i}",
                    )

    # ── Construction du commentaire ───────────────────────────────────────────

    _MSG = {
        "Progression 0/0":         ("PROGRESSION NULLE",
                                    "Rouvrez l'enquete et traitez toutes les sections."),
        "Progression trop faible": ("PROGRESSION INSUFFISANTE",
                                    "Moins de 50 questions traitees. Completez toutes les sections."),
        "inaccessibles":           ("DONNEES INACCESSIBLES",
                                    "Les donnees n'ont pu etre lues. Contactez le superviseur."),
        "Horodatage manquant":     ("HORODATAGE MANQUANT",
                                    "Les heures de debut ou de fin sont absentes. Verifiez date et end_interview."),
        "Duree insuffisante":      ("DUREE INSUFFISANTE",
                                    "Duree inferieure a 45 min. Reconduisez l'enquete completement."),
        "Coordonnees GPS":         ("GPS MANQUANT",
                                    "Les coordonnees GPS du menage sont absentes. Activez le GPS et re-saisissez."),
        "Precision GPS":           ("PRECISION GPS INSUFFISANTE",
                                    "Precision superieure a 20 m. Attendez un meilleur signal avant de valider."),
        "Consentement":            ("CONSENTEMENT MANQUANT",
                                    "Le repondant n'a pas donne son consentement. Verifiez la question consent."),
        "Contradiction":           ("DECLARATION CONTRADICTOIRE",
                                    "Vous avez declare des enfants (Oui) mais aucun dossier enfant n'est rempli."),
        "Age manquant":            ("AGE ENFANT MANQUANT",
                                    "L'age d'un ou plusieurs enfants est absent. Verifiez les dates de naissance."),
        "pour un enfant interviewe":   ("DATE NAISSANCE ENFANT INTERVIEWE",
                                    "La date de naissance d'un enfant interviewe est absente ou invalide. Verifiez child_year_birth_new."),
        "pour un nouvel enfant":        ("DATE NAISSANCE NOUVEL ENFANT",
                                    "La date de naissance d'un nouvel enfant est absente ou invalide. Verifiez new_child_year_birth."),
        "pour un enfant":               ("DATE NAISSANCE ENFANT PRECHARGE",
                                    "La date de naissance d'un enfant pre-charge esquive est absente ou invalide. Verifiez child_year_birth_new."),
        "esquivé sans justification":   ("ENFANT ESQUIVE SANS JUSTIFICATION",
                                    "Un enfant pre-enregistre a ete esquive sans etre declare comme doublon. Interviewez l'enfant ou signalez le doublon."),
        "déclaré doublon":              ("DOUBLON SANS NOM",
                                    "Un enfant est declare doublon mais le nom du doublon est vide. Selectionnez le nom dans duplicates_child_name."),
        "Fraude suspectée":            ("ENFANTS NON INTERVIEWS",
                                    "Des enfants etaient pre-enregistres mais tous ont ete refuses. Completez les dossiers enfants."),

        "Nom d'enfant invalide":   ("NOM ENFANT INVALIDE",
                                    "Le nom d'un enfant contient uniquement des chiffres. Saisissez le vrai nom."),
        "Formatage":               ("FORMATAGE INCORRECT",
                                    "Des champs sont en minuscules ou avec accents. Tout saisir en MAJUSCULES sans accents."),
        "Impossible d'extraire":   ("DONNEES INACCESSIBLES",
                                    "Les donnees n'ont pu etre lues. Contactez le superviseur."),
        "Impossible d'obtenir":    ("DONNEES INACCESSIBLES",
                                    "Les donnees n'ont pu etre lues. Contactez le superviseur."),
    }

    def _build_comment(self) -> None:
        if not self.result.errors:
            self.result.decision = "HQ_APPROVE"
            self.result.comment  = (
                "ENQUETE APPROUVEE PAR HQ\n"
                f"Enqueteur : {self.result.enumerator}\n"
                f"Progression : {self.result.progress_str}\n"
                f"Enfants traites : {self.result.children_count}"
            )
            return

        self.result.decision = "HQ_REJECT"

        seen_titres = set()
        lignes = [
            f"ENQUETE REJETEE - {len(self.result.errors)} anomalie(s)",
            "",
        ]

        num = 1
        for err in self.result.errors:
            titre  = "ANOMALIE"
            action = err
            for trigger, (t, a) in self._MSG.items():
                if trigger in err:
                    titre  = t
                    action = a
                    break
            if titre in seen_titres:
                continue
            seen_titres.add(titre)
            lignes.append(f"{num}. {titre}")
            lignes.append(f"   -> {action}")
            lignes.append("")
            num += 1

        lignes.append("Corrigez les points ci-dessus et renvoyez l'enquete.")
        self.result.comment = "\n".join(lignes)

# ==============================================================================
# 12. EXÉCUTION PRINCIPALE
# ==============================================================================

def executer_validation_hq(
    mode_test: bool          = False,
    mode_diag: bool          = False,
    limite:    Optional[int] = None,
) -> None:
    logger.info("=" * 70)
    logger.info(
        f"DEMARRAGE CLMRS HQ v11.9 — "
        f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )
    mode_label = (
        "DIAGNOSTIC"              if mode_diag
        else "TEST (lecture seule)" if mode_test
        else "PRODUCTION HQ"
    )
    logger.info(f"Mode : {mode_label}")
    logger.info("=" * 70)

    client = SurveySolutionsHQClient()
    if not client.test_connection():
        logger.error("Abandon — serveur inaccessible ou droits insuffisants.")
        return

    try:
        entretiens = client.get_interviews(limit=limite)
    except Exception as e:
        logger.error(f"Erreur fatale lors de la récupération : {e}")
        return

    if not entretiens:
        logger.info("Aucune enquête à traiter.")
        return

    total = len(entretiens)
    logger.info(f"Enquêtes récupérées : {total}")

    if mode_diag:
        client.run_diagnostic(entretiens, nb=min(5, total))
        return

    logger.info("=" * 70)
    done_ids  = load_checkpoint()
    stats_run = {
        "total":   total,
        "approuves": 0,
        "rejetes":   0,
        "erreurs":   0,
        "sautes":    0,
        "erreurs_detail": [],
    }
    results_export = []

    for index, inter in enumerate(entretiens, 1):
        i_id   = safe_get(inter, "InterviewId", default="")
        i_key  = safe_get(inter, "Key",         default="?")
        statut = safe_get(inter, "Status",      default="?")

        if not i_id:
            logger.error(f"[{index}/{total}] InterviewId manquant — ignoré.")
            stats_run["erreurs"] += 1
            continue

        if i_id in done_ids:
            logger.info(f"[{index}/{total}] {i_key} ({i_id}) — déjà traité (checkpoint), sauté.")
            stats_run["sautes"] += 1
            continue

        logger.info(
            f"\n[{index}/{total}] {i_key} ({i_id}) | "
            f"Statut : {statut} | "
            f"Enquêteur : {safe_get(inter, 'ResponsibleName', default='?')}"
        )

        try:
            detail = client.get_interview_details(i_id)
            istats = client.get_interview_stats(i_id)

            if not detail:
                msg = (
                    f"Détails inaccessibles (statut={statut}). "
                    "Lancez --diagnostic pour identifier la cause HTTP."
                )
                logger.error(f"  {msg}")
                stats_run["erreurs"] += 1
                stats_run["erreurs_detail"].append((i_id, msg))
                continue

            validator = HQInterviewValidator(inter, detail, istats)
            res       = validator.validate()
            results_export.append(res)

            if res.decision == "HQ_APPROVE":
                logger.info(
                    f"  HQ_APPROVE — {res.progress_str} | "
                    f"Enfants : {res.children_count}"
                )
                if not mode_test:
                    ok, err_msg = client.hq_approve(i_id, res.comment)
                    if ok:
                        stats_run["approuves"] += 1
                    else:
                        logger.error(f"  hqapprove échoué : {err_msg}")
                        stats_run["erreurs"] += 1
                        stats_run["erreurs_detail"].append((i_id, f"hqapprove: {err_msg}"))
                        continue  # Ne pas marquer comme traité
                else:
                    stats_run["approuves"] += 1

            else:
                logger.warning(
                    f"  HQ_REJECT — {len(res.errors)} anomalie(s) | {res.progress_str}"
                )
                for err in res.errors:
                    logger.warning(f"    • {err}")

                if not mode_test:
                    if statut == "RejectedBySupervisor":
                        # Seul /hqreject accepte ce statut
                        ok, err_msg = client.hq_reject(i_id, res.comment)
                        if ok:
                            stats_run["rejetes"] += 1
                        else:
                            logger.error(f"  hqreject échoué : {err_msg}")
                            stats_run["erreurs"] += 1
                            stats_run["erreurs_detail"].append((i_id, f"hqreject: {err_msg}"))
                            continue

                    elif statut == "ApprovedBySupervisor":
                        # Unapprove d'abord, puis reject superviseur
                        ok_un, err_un = client.supervisor_unapprove(i_id)
                        if not ok_un:
                            logger.error(f"  unapprove échoué : {err_un}")
                            stats_run["erreurs"] += 1
                            stats_run["erreurs_detail"].append((i_id, f"unapprove: {err_un}"))
                            continue
                        logger.info("  unapprove OK")
                        time.sleep(0.5)
                        ok, err_msg = client.supervisor_reject(i_id, res.comment)
                        if ok:
                            stats_run["rejetes"] += 1
                        else:
                            logger.error(f"  reject échoué après unapprove : {err_msg}")
                            stats_run["erreurs"] += 1
                            stats_run["erreurs_detail"].append((i_id, f"reject: {err_msg}"))
                            continue

                    else:
                        # Completed → reject superviseur direct
                        ok, err_msg = client.supervisor_reject(i_id, res.comment)
                        if ok:
                            stats_run["rejetes"] += 1
                        else:
                            logger.error(f"  reject échoué : {err_msg}")
                            stats_run["erreurs"] += 1
                            stats_run["erreurs_detail"].append((i_id, f"reject: {err_msg}"))
                            continue
                else:
                    stats_run["rejetes"] += 1

            done_ids.add(i_id)
            save_checkpoint(done_ids)

        except Exception as e:
            msg = f"{type(e).__name__}: {str(e)[:300]}"
            logger.error(f"  ERREUR SYSTÈME pour {i_id} : {msg}")
            stats_run["erreurs"] += 1
            stats_run["erreurs_detail"].append((i_id, msg))

        time.sleep(CONFIG["sleep_between_calls"])

    # ── Statistiques finales ──────────────────────────────────────────────────
    traites = total - stats_run["sautes"]
    logger.info("\n" + "=" * 70)
    logger.info("STATISTIQUES FINALES")
    logger.info("=" * 70)
    logger.info(f"Total récupéré      : {total}")
    logger.info(f"Sautés (checkpoint) : {stats_run['sautes']}")
    logger.info(f"Traités ce run      : {traites}")
    logger.info(
        f"HQ Approuvés        : {stats_run['approuves']} "
        f"({stats_run['approuves'] * 100 / max(1, traites):.1f}%)"
    )
    logger.info(
        f"Rejetés             : {stats_run['rejetes']} "
        f"({stats_run['rejetes'] * 100 / max(1, traites):.1f}%)"
    )
    logger.info(f"Erreurs système     : {stats_run['erreurs']}")

    if stats_run["erreurs_detail"]:
        logger.warning("\nDÉTAIL ERREURS (10 premières) :")
        for i_id, msg in stats_run["erreurs_detail"][:10]:
            logger.warning(f"  {i_id} → {msg}")
        reste = len(stats_run["erreurs_detail"]) - 10
        if reste > 0:
            logger.warning(f"  … et {reste} autre(s). Lancez --diagnostic pour analyser.")

    logger.info("=" * 70)

    if CONFIG["export_report"] and results_export:
        export_excel(results_export, CONFIG["report_file"])

    if not limite and stats_run["erreurs"] == 0:
        clear_checkpoint()

# ==============================================================================
# 13. POINT D'ENTRÉE
# ==============================================================================

if __name__ == "__main__":
    args      = sys.argv[1:]
    mode_test = "--test"       in args or "-t" in args
    mode_diag = "--diagnostic" in args or "-d" in args
    reset     = "--reset"      in args

    if reset:
        clear_checkpoint()
        logger.info("Checkpoint réinitialisé. Relancez sans --reset.")
        sys.exit(0)

    limite: Optional[int] = None
    for arg in args:
        if arg.startswith("--limit="):
            try:
                limite = int(arg.split("=", 1)[1])
            except ValueError:
                logger.warning(f"Valeur --limit invalide : {arg}")

    try:
        executer_validation_hq(
            mode_test=mode_test,
            mode_diag=mode_diag,
            limite=limite,
        )
    except KeyboardInterrupt:
        logger.info("Interruption manuelle — checkpoint sauvegardé.")
        sys.exit(0)
    except Exception as e:
        logger.critical(f"Erreur fatale non gérée : {e}")
        sys.exit(1)
